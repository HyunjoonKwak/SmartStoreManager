import os
from openpyxl import load_workbook
import shutil
import time
import datetime
import pandas as pd
import win32com.client

def read_option():
    df_option = pd.read_csv('./사전작업/ople_option.txt', header=None)
    df_option.columns = ['option']
    path = df_option['option'][0]
    fname = df_option['option'][1]
    start = int(df_option['option'][2])
    paging = int(df_option['option'][3])
    up_margin = int(df_option['option'][4])
    dn_margin = int(df_option['option'][5])
    return path, fname, start, paging, up_margin, dn_margin

tpath, fname, start, paging, up_margin, dn_margin = read_option()
list_file = tpath+fname

up = float(up_margin)
dn = float(dn_margin)
print(f"기준할인율 : up:{up} %, dn:{dn}%")

dt_today = datetime.date.today()
dt_now = datetime.datetime.now()
date = dt_now.strftime('%Y%m%d%H%M')
today = dt_today.strftime('%Y%m%d')
output = './사전작업/output_Ople'+today+'.xlsx'                                  # output_20230207.xlsx
plist = './오플가격체크결과/'+fname.rstrip('.xlsx') + '_'+ today + '.xlsx'      # data/상품업로드목록_Latest_20230207.xlsx
slist = './오플가격체크결과/스마트스토어상품_Ople_'+today+'.xlsx'                     # data/스마트스토어상품_20230207.xlsx
shutil.copy(list_file, plist)                                               # option.txt 에서 지정한 원본 업로드목록 을 복사 한다.
shutil.copy2('./사전작업/스마트스토어상품_오플_'+today+'.xlsx', slist)            # 다운받은 상품업로드용 목록 을 복사 한다.

wa = load_workbook(output)            # output_20230207.xlsx
waa = wa.active
# print(f"wa max row is {waa.max_row}, max column is {waa.max_column}")

wb = load_workbook(plist)        # data/상품업로드목록_Latest_20230207.xlsx
wbb = wb.worksheets[0]
# print(f"wb max row is {wbb.max_row}, max column is {wbb.max_column}")

wc = load_workbook(slist)       # data/스마트스토어상품_20230207.xlsx
wcc = wc.active
# print(f"wc max row is {wcc.max_row}, max column is {wcc.max_column}")
sold_out_count = 0
new_soldout = 0
fin_soldout = 0
for j in range(start+2,paging+2):
    if waa.cell(j,9).value is not None:     # 현재의 가격이 None 이 아닐경우에만 실행
        for i in range(9,11):
            wbb.cell(row=j, column = i+3, value = ' ')
            wbb.cell(row=j, column = i+3, value = waa.cell(j,i).value)  # 쿠팡에서 얻어온 현재 가격과 품절상태를 쿠팡가와 품절에 업데이트

        cp_soldout = wbb.cell(j,13).value # 쿠팡 품절상태
        ss_soldout = wbb.cell(j,11).value # 위더스 품절상태
        # print(f"{j-1}: cp {cp_soldout}, ss {ss_soldout}")
        if cp_soldout == 'x' and ss_soldout == 'empty':
            # print(f"{j-1} 품절상태 동일")
            sold_out_count += 1
        else:
            if cp_soldout == 'x':
                print(f"{j-1} 신규품절. 상품이름: {wbb.cell(j,9).value} link: {wbb.cell(j,5).value}")
                wbb.cell(j,11).value = 'empty'
                sold_out_count += 1
                new_soldout += 1             
            elif ss_soldout == 'empty':
                print(f"{j-1} 품절풀림 상품이름: {wbb.cell(j,9).value} link: {wbb.cell(j,5).value}")
                wbb.cell(j,11).value = ' '
                fin_soldout += 1

print(f"전체 품절항목 : {sold_out_count}, 신규품절 {new_soldout}항목 , 품절풀림 {fin_soldout}항목")                            

price_up = 0
price_dn = 0
for i in range(wbb.max_row):
    now_price = wbb.cell(i+2, 14).value
    if now_price is not None:
        # print(f"now_price : {now_price}, {type(now_price)}")
        curr_price = wbb.cell(i+2, 12).value
        if curr_price is not None:
            price_gap = curr_price - now_price          # 가격차이
            percent = float(price_gap / now_price)      # 인상율/인하율
            percent *= 100
            percent = round(percent, 2)
            price_update = False
            # print(f"쿠팡현재가 : {curr_price}, 스스 현재가 : {now_price}, 갭 :{price_gap},{percent}% ")
            if(price_gap != 0):
                if(price_gap > 0 ):
                    if(percent >= up):
                        price_update = True
                        price_up += 1
                elif(price_gap < 0):
                    percent = -percent
                    if(percent >= dn):
                        price_update = True
                        price_dn += 1

        if(price_update):
            wbb.cell(row=i+2, column = 14, value = ' ' )
            wbb.cell(row=i+2, column = 14, value = curr_price)

print(f" {price_up}항목 인상, {price_dn}항목 인하")

df_slist = pd.read_excel(slist, keep_default_na=False)
df_slist = df_slist.fillna('')

for i, prod_num in enumerate(df_slist.iloc[4:,0]):
    prod_num = int(prod_num)
    wbb.cell(row=i+2, column = 25, value = ' ' )
    wbb.cell(row=i+2, column = 25, value = prod_num)
old_file = plist
plist = plist.rstrip('.xlsx')+ '_v'+'.xlsx'
wb.save(plist)      # _v 붙어있는 파일로 저장. 값으로 변환하기 위해서.
os.remove(old_file)

#win32com
# _path = "C:\\Users\\hyunj\\내 드라이브\\파이썬 코드\\상품현재가격 확인\\"
_path = "C:\\coding\\상품현재가격 확인_ver2.0\\"
excel = win32com.client.Dispatch("Excel.Application")   #excel 사용할 수 있게 설정
_path = _path + plist
temp_wb = excel.Workbooks.Open(_path)                   #임시 Workbook 객체 생성 및 엑셀 열기
temp_wb.Save()                                          #저장
excel.quit()                                            #excel 종료

wb = load_workbook(plist, data_only=True)
wbb = wb.active
wb.save(plist)                 # 수식을 값으로 변환해서 저장.

# 이걸 다시 스마트스토어상품 엑셀로 옮기고
# plist 의 ''상품가격' 탭과 '할인가'탭을
# slist 의 '판매가' = 6번쨰 탭과 'PC즉시할인값'=(AY= 51), '모바일즉시할인값'=(BA=53)에 복사한다.
baesong_template = wbb.cell(2,31).value
for i in range(wcc.max_row):
    fprice = wbb.cell(i+2,27).value
    sprice = wbb.cell(i+2,28).value      # 배송비 2,500원 녹이기
    if fprice is not None:
        # print(f"f price : {fprice} , sale {sprice}, {type(sprice)}")
        # if type(sprice) == 'int' and sprice > 2500: sprice -= 2500   # 배송비 2500원 녹이기
        wcc.cell(row = i+6, column = 6, value = ' ')
        wcc.cell(row = i+6, column = 6, value = fprice)
        wcc.cell(row = i+6, column = 30, value = baesong_template)
        wcc.cell(row = i+6, column = 53, value = ' ')
        wcc.cell(row = i+6, column = 51, value = sprice)
        wcc.cell(row = i+6, column = 51, value = sprice)
        wcc.cell(row = i+6, column = 53, value = sprice)

wc.save(slist)
# wc.save('./data/test2.xlsx')

if False:
    # 파일을 하나 더 복사한 다음
    front = slist.rstrip('.xlsx')+ '_1'+'.xlsx'
    backend = slist.rstrip('.xlsx')+ '_2'+'.xlsx'
    shutil.copy(slist, front)
    shutil.copy(slist, backend)

    os.remove(slist)


    print(f"{wcc.max_row}")

    wf = load_workbook(front)
    wff = wf.active
    we = load_workbook(backend)
    wee = we.active

    # 500개 단위로 자른다.
    wff.delete_rows(3,3)
    wff.delete_rows(500,500)
    wee.delete_rows(3, 500)

    wf.save(front)
    we.save(backend)
else:
    tcount = wcc.max_row
    lcount = int(tcount / 500 + 1)
    print(f"total : {tcount}, loop : {lcount}")

    wf = load_workbook(slist)
    wff = wf.active
    wff.delete_rows(3,3)
    wf.save(slist)

    for i in range(lcount):
    # 파일을 하나 더 복사한 다음
        file = slist.rstrip('.xlsx')+ '_'+str(i)+'.xlsx'
        shutil.copy(slist, file)
        wf = load_workbook(file)
        wff = wf.active
        # 500개 단위로 자른다.
        j = i + 1
        until = tcount - 500*j
        amount = i*500
        split = 500 
        if(amount > 0):
            wff.delete_rows(3,amount-3*i)
            print(f"delete 3~{amount-3*i}")
        wff.delete_rows(split,until)
        print(f"delete {split} ~ {until}")
        wf.save(file)
    
    os.remove(slist)