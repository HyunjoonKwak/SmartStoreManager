from openpyxl import load_workbook
import pandas as pd
import os

file = "./price_format_template.xlsx"
wb = load_workbook(file)
ws = wb.active

base = "./data_crawl/prod_result_copy.xlsx"
wa = load_workbook(base)
waa = wa.active

for j in range(1,3):
    for i in range(14,20):
        print(i, ',',ws.cell(j,i).value)
        waa.cell(row=j, column = i, value = ws.cell(j,i).value)

wa.save(base)