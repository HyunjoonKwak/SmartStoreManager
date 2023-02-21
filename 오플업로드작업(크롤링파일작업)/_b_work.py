import time
import datetime
import glob
## 데이터 라이브러리
import pandas as pd
import os
from os import rename
import shutil
from distutils.dir_util import copy_tree
from openpyxl import load_workbook
from pathlib import Path

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)
 
# 이미지 파일 이름 변경
def ImageRename(path):
    changeimagedir = './data_crawl/renamed'

    # shutil.rmtree(changeimagedir)

    copy_tree(path, changeimagedir)

    file_list = os.listdir(path)

    target = changeimagedir+'/'
    prefix = 'xxx_'
    for name in file_list:
        os.rename(target + name, target + prefix + name)

    for name in file_list:
        file_name = os.path.splitext(name)[0]

        i = 0       
        for i, rd in enumerate(df1["Idx"]):
            if int(file_name) == rd:
                print("filename ",file_name , "Prd_ID", df1["Prd_ID"][i])
                f = df1["Prd_ID"][i]
                print(f)
                try:
                    os.rename(target + prefix + file_name+'.jpg', target + str(f)+'.jpg')
                except:
                    os.rename(target + prefix + file_name+'.png', target + str(f)+'.jpg')

#사용하지 않는 파일들은 삭제 해줌
def deleteImage(path):
    file_list = glob.glob(f"{path}/xxx_*.jpg")
    for f in file_list:
        os.remove(f)
    file_list = glob.glob(f"{path}/xxx_*.png")
    for f in file_list:
        os.remove(f)

# 50 개씩 짤라서 폴더이동    
def FileNameCheck(path, flist, folder_slicecount):  
    fcount = 0
    path = Path(path)
    file_list = path.iterdir()
    for item in file_list:
        if item.is_file():
            # print(item.name)
            flist.append(item.name)
            fcount += 1
    
    # folder_slicecount = 5
    
    folderCount = fcount / folder_slicecount
    if fcount > folder_slicecount:
        for i in range(int(folderCount)+1):
            # print(f"{i+1} 이름 폴더를 만들고")
            createFolder('./data_crawl/renamed'+'/'+str(i+1))
        
        for i in range(fcount):
            k = int(i/folder_slicecount)
            # print(f"move file {flist[i]} 를 {k+1} 폴더로 옮긴다.")
            shutil.move('./data_crawl/renamed'+'/'+flist[i], './data_crawl/renamed'+'/'+str(k+1))

# 판매상품 관리용 포맷으로 변경
def make_prod_list(list, prod):
    df_prod['구분'] = list['Src']
    df_prod['code'] = list['code']
    df_prod['링크'] = list['링크']
    df_prod['Brand'] = list['Brand']
    df_prod['Title'] = list['Title']
    df_prod['Prd_ID'] = list['Prd_ID']
    df_prod['Soldout'] = list['SoldOut']
    df_prod['Now'] = list['Now']
    df_prod['현재가격'] = list['Now']
    t = df_prod['기준'].copy()
    t[0] = '15%'
    t[1] = 20000
    df_prod['기준'] = t
    for i, val in enumerate(list['Idx']):
        n = df_prod['NO'].copy()
        n[i] = i+2
        df_prod['NO'] = n
    return

# 옵션정보 읽어오기
def read_option():
    option = []
    f = open("./사전작업/worktype.txt", "r")

    while True:
        line = f.readline().strip()
        if not line: break
        option.append(line)

    for i, ID in enumerate(option):
        if(i == 0):
            self_crawl = ID
        else:
            test = ID

    if(self_crawl == 'True'): self_crawl = True
    else: self_crawl = False
    if(test == 'True'): test = True
    else: test = False

    return self_crawl, test
###############################################################################################################
#   <MAIN
###############################################################################################################
###############################################################################################################
#   self_crawl 설정
#   제공된 폼 이용 = False,  자체 크롤링 폼 이용 = True
#   True : Prework 작업 선결할 것
#   False : OpleMakeForm.xlsx 작업 선결할 것
###############################################################################################################
sheetname = 'ople'

self_crawl, test = read_option()
print("self_crawl : ", self_crawl, " test : ", test)

if self_crawl == True:
    list_file = "./data_collect/result_first.xlsx"
else :    
    list_file = "./data_crawl/result_first.xlsx"

df1 = pd.read_excel(list_file, keep_default_na=False)
df1 = df1.fillna('')

flist = []
if self_crawl == False:
    ImageRename('./사전작업/image')
    deleteImage('./data_crawl/renamed')
    # FileNameCheck('./data_crawl/renamed', flist, 50)

#열 이름 설정
df_prod = pd.DataFrame(columns = ['구분','NO','code','상품번호','링크','Brand','Title','Prd_ID','Soldout','Now','기준','품절','현재가격','15%','표시가격'])
make_prod_list(df1, df_prod)
# print(df_prod)
dt_today = datetime.date.today()
dt_now = datetime.datetime.now()
date = dt_now.strftime('%Y%m%d%H%M')

if self_crawl == True:
    if(test):
        path = './data_collect/prod_result_'+date+'.xlsx'
        sheetname += date
    else:
        path = './data_collect/prod_result_ready'+'.xlsx'
else:
    if(test):
        path = './data_crawl/prod_result_'+date+'.xlsx'
        sheetname += date
    else:
       path = './data_crawl/prod_result_ready'+'.xlsx'

with pd.ExcelWriter(path) as writer:
    df_prod.to_excel(writer, sheet_name =sheetname, index=False)

file = "./사전작업/price_format_template.xlsx"
wb = load_workbook(file)
ws = wb.active

if self_crawl == True:
    base = "./data_collect/prod_result_ready.xlsx"
else :
    base = "./data_crawl/prod_result_ready.xlsx"
wa = load_workbook(base)
waa = wa.active

for j in range(1,3):
    for i in range(14,20):
        # print(i, ',',ws.cell(j,i).value)
        waa.cell(row=j, column = i, value = ws.cell(j,i).value)

wa.save(base)