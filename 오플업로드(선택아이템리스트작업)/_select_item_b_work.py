import time
import datetime
## 데이터 라이브러리
import pandas as pd
import os
from os import rename
import shutil
from distutils.dir_util import copy_tree
from openpyxl import load_workbook

# 판매상품 관리용 포맷으로 변경
def make_prod_list(list, prod):
    df_prod['구분'] = list['Src']
    df_prod['code'] = list['code']
    df_prod['링크'] = list['링크']
    df_prod['Brand'] = list['Brand']
    df_prod['Title'] = list['Title']
    df_prod['Prd_ID'] = list['Prd_ID']
    df_prod['Soldout'] = list['SoldOut']
    df_prod['Now'] = list['Now']
    df_prod['현재가격'] = list['Now']
    t = df_prod['기준'].copy()
    t[0] = '15%'
    t[1] = 20000
    df_prod['기준'] = t
    for i, val in enumerate(list['Idx']):
        n = df_prod['NO'].copy()
        n[i] = i+2
        df_prod['NO'] = n
    return

# 옵션정보 읽어오기
def read_option():
    option = []
    f = open("./사전작업/worktype.txt", "r")

    while True:
        line = f.readline().strip()
        if not line: break
        option.append(line)

    for i, ID in enumerate(option):
        if(i == 0):
            self_crawl = ID
        else:
            test = ID

    if(self_crawl == 'True'): self_crawl = True
    else: self_crawl = False
    if(test == 'True'): test = True
    else: test = False

    return self_crawl, test

###############################################################################################################
#   <MAIN
###############################################################################################################
self_crawl, test = read_option()
print(" test : ", test)

list_file = "./data_select/result_first.xlsx"

df1 = pd.read_excel(list_file, keep_default_na=False)
df1 = df1.fillna('')

#열 이름 설정
df_prod = pd.DataFrame(columns = ['구분','NO','code','상품번호','링크','Brand','Title','Prd_ID','Soldout','Now','기준','품절','현재가격','15%','표시가격'])
make_prod_list(df1, df_prod)
# print(df_prod)
dt_today = datetime.date.today()
dt_now = datetime.datetime.now()
date = dt_now.strftime('%Y%m%d%H%M')

if(test):
    path = './data_select/prod_result_'+date+'.xlsx'
else:
    path = './data_select/prod_result_ready'+'.xlsx'
with pd.ExcelWriter(path) as writer:
    df_prod.to_excel(writer, sheet_name ='Sheet1', index=False)

file = "./사전작업/price_format_template.xlsx"
wb = load_workbook(file)
ws = wb.active

base = "./data_select/prod_result_ready.xlsx"
wa = load_workbook(base)
waa = wa.active

for j in range(1,3):
    for i in range(14,20):
        # print(i, ',',ws.cell(j,i).value)
        waa.cell(row=j, column = i, value = ws.cell(j,i).value)

wa.save(base)